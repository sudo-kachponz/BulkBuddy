import os
import uuid
import pandas as pd
from fpdf import FPDF
from fastapi import APIRouter
from pydantic import BaseModel
from typing import List, Dict, Any
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

router = APIRouter(prefix="/api")

class ReportRequest(BaseModel):
    data: List[Dict[str, Any]]
    send_email: bool = False
    to_email: str = "neutracksudo@gmail.com"

class UpdateSheetRequest(BaseModel):
    sheet_name: str
    data: List[Dict[str, Any]]
    send_email: bool = False
    to_email: str = "neutracksudo@gmail.com"

def send_email_with_attachments(to_email: str, subject: str, html_body: str, file_paths: List[str]):
    smtp_user = os.environ.get("SMTP_USERNAME")
    smtp_pass = os.environ.get("SMTP_PASSWORD")

    if not smtp_user or not smtp_pass:
        raise Exception("SMTP credentials (SMTP_USERNAME, SMTP_PASSWORD) are missing in .env")

    msg = MIMEMultipart()
    msg['From'] = smtp_user
    msg['To'] = to_email
    msg['Subject'] = subject

    msg.attach(MIMEText(html_body, 'html'))

    for file_path in file_paths:
        if os.path.exists(file_path):
            with open(file_path, "rb") as f:
                part = MIMEApplication(f.read(), Name=os.path.basename(file_path))
                part['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
                msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(smtp_user, smtp_pass)
        server.send_message(msg)

@router.post("/update-sheet")
async def update_sheet_api(request: UpdateSheetRequest):
    try:
        columns_order = [
            "nama", "gelarsbl", "gelarsdh", "kelamin", "tgl_lhr", "kota_lhr", 
            "warga_negara", "no_ktp", "kota_ktp", "exp_ktp", "jenis_id_tambahan", 
            "no_id_tambahan", "ibu_kandung", "sts_kawin", "alamat1", "alamat2", 
            "kodepos", "telp_rumah", "handphone", "email", "pekerjaan", "jabatan", 
            "employer_name", "kode_industri", "tgl_mulai", "gaji", "pen_lain", 
            "cif_no", "currency", "produk", "biaya_admin", "tujuan_buka", 
            "kode_cabang", "bansos_type", "consent"
        ]
        
        sheet_values = []
        for row in request.data:
            formatted_row = [str(row.get(col, "")) for col in columns_order]
            sheet_values.append(formatted_row)

        try:
            import gspread
        except ImportError:
            raise Exception("Library 'gspread' tidak ditemukan. Jalankan: pip install gspread google-auth")

        # Pastikan file credentials.json (Service Account) ada di root folder
        creds_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), 'credentials.json')
        if not os.path.exists(creds_path):
            raise Exception(f"File kredensial Service Account tidak ditemukan di {creds_path}. Silakan unduh dari Google Cloud Console dan simpan sebagai 'credentials.json' di folder root BulkBuddy.")

        gc = gspread.service_account(filename=creds_path)
        spreadsheet = gc.open(request.sheet_name)
        worksheet = spreadsheet.sheet1
        
        # Hapus data lama (mulai dari baris 2 ke bawah, menyisakan header) dan timpa
        worksheet.batch_clear(["A2:AI"]) 
        
        if sheet_values:
            worksheet.update(values=sheet_values, range_name='A2')

        return {"status": "success", "message": f"Data saved to {request.sheet_name} instantly!"}

    except Exception as e:
        print(f"Error updating sheet: {type(e).__name__} - {e}")
        from fastapi import HTTPException
        if "SpreadsheetNotFound" in str(type(e).__name__):
            raise HTTPException(status_code=404, detail="Spreadsheet tidak ditemukan oleh Service Account. Pastikan file sudah di-share ke email bulkbuddy@animated-vector-434910-g3.iam.gserviceaccount.com")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/generate-reports")
async def generate_reports(req: ReportRequest):
    data = req.data
    if not data:
        return {"error": "No data provided"}

    # Prevent Excel from dropping leading zeros or using scientific notation
    target_columns = ["No Telp Rumah", "No. Handphone", "NO KTP / PASSPORT", "NO IDENTITAS TAMBAHAN", "KODEPOS", "CIF_NO"]
    for row in data:
        # Add template default values
        row["KERJAPSW"] = row.get("KERJAPSW", "")
        row["WARGA NEGARA"] = row.get("WARGA NEGARA", "000")
        row["EMPLOYER NAME"] = row.get("EMPLOYER NAME", "PT SUTET")
        row["KODE_INDUSTRI"] = row.get("KODE_INDUSTRI", "09")
        from datetime import datetime
        row["TGL_MULA"] = row.get("TGL_MULA", datetime.now().strftime("%d%m%Y"))
        row["GAJI"] = row.get("GAJI", "300000")
        row["PEN_LAIN"] = row.get("PEN_LAIN", "0")
        row["CIF_NO"] = row.get("CIF_NO", "0")
        row["CURRENCY"] = row.get("CURRENCY", "IDR")
        row["PRODUK"] = row.get("PRODUK", "TABMANDIRI")
        row["BIAYA ADMIN KHU"] = row.get("BIAYA ADMIN KHU", "")
        row["TUJUAN BUKA REKENING"] = row.get("TUJUAN BUKA REKENING", "A")
        row["KODE CABANG"] = row.get("KODE CABANG", "12000")
        row["BANSOS TYPE"] = row.get("BANSOS TYPE", "")
        row["CONSENT"] = row.get("CONSENT", "YYYY")

        for col in target_columns:
            if col in row and row[col]:
                val_str = str(row[col]).strip()
                if val_str.isdigit():
                    row[col] = f"'{val_str}"

    unique_id = str(uuid.uuid4())[:8]
    excel_path = f"/tmp/laporan_nasabah_{unique_id}.xlsx"
    pdf_path = f"/tmp/laporan_nasabah_{unique_id}.pdf"

    # 1. Generate Excel using Pandas
    df = pd.DataFrame(data)
    try:
        from openpyxl.styles import Font, PatternFill, Border, Side
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            worksheet = writer.sheets['Sheet1']
            
            header_font = Font(name='Arial', size=16, bold=True)
            header_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # Style header row
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                
            # Resize columns and apply borders to all cells
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    cell.border = thin_border
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                worksheet.column_dimensions[column].width = max_length + 3
    except ImportError:
        df.to_excel(excel_path, index=False)

    # 2. Generate PDF using FPDF
    class PDF(FPDF):
        def header(self):
            self.set_font("helvetica", "B", 12)
            self.cell(0, 10, "Laporan Data Nasabah (BulkBuddy)", align="C")
            self.ln(15)

    pdf = PDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    
    columns = ["nama", "no_ktp", "kelamin", "tgl_lhr", "handphone", "produk", "kode_cabang"]
    col_widths = [45, 45, 15, 25, 35, 30, 25]
    
    # Table Header
    pdf.set_font("helvetica", "B", 8)
    for i, col in enumerate(columns):
        pdf.cell(col_widths[i], 8, str(col).upper(), border=1, align="C")
    pdf.ln()

    # Table Rows
    pdf.set_font("helvetica", "", 8)
    for row in data:
        for i, col in enumerate(columns):
            val = str(row.get(col, ""))[:25]
            pdf.cell(col_widths[i], 8, val, border=1)
        pdf.ln()

    pdf.output(pdf_path)
    
    if req.send_email:
        subject = "[Permohonan Pembukaan Rekening BULK Tabungan Reguler - PLN SUTET]"
        html_body = f"""
        <b>Cash & Trade Operations Group</b><br>
        <b>Bulk Payment & Account Opening Department</b><br>
        Sentra Mandiri Gedung B Lt. 4<br>
        JL. RP Soeroso No. 2-4<br>
        Jakarta 10330<br><br>
        ---<br><br>
        <b>Perihal:</b> : <span style="text-decoration: underline;"><b>[Permohonan Pembukaan Rekening BULK Tabungan Reguler]</b></span><br><br>
        Sehubungan dengan diadakannya kerjasama pembukaan Tabungan Reguler antara PT Sutet... dengan Bank Mandiri Tanjung Priok Enggano (12000), dengan ini kami sampaikan permintaan pembukaan rekening secara bulk untuk dapat diproses sesuai informasi sebagai berikut:<br><br>
        <ul>
        <li><b>Jumlah Rekening:</b> : <b>[{len(data)} Rekening (rincian terlampir)]</b></li>
        <li><b>Jenis:</b> : ACTIVE</li>
        <li><b>Kode Tabungan:</b> : TABMANDIRI</li>
        </ul><br>
        Data dimaksud telah kami periksa dan diyakini kebenarannya telah sesuai e-KTP, yaitu :<br><br>
        <table border="1" cellpadding="5" cellspacing="0" style="border-collapse: collapse;">
          <tr>
            <th>Field</th>
            <th>Keterangan</th>
          </tr>
          <tr>
            <td>NIK</td>
            <td>Wajib 16 Digit (sesuai e-KTP)</td>
          </tr>
          <tr>
            <td>Nama</td>
            <td>Ejaan/ spasi (sama persis dengan e-KTP)</td>
          </tr>
          <tr>
            <td>Tanggal Lahir</td>
            <td>Tanggal Bulan dan Tahun (sama persis dengan e-KTP)</td>
          </tr>
        </table><br>
        Apabila terdapat kesalahan data (tidak sesuai e-KTP), segala risiko dan akibat yang timbul setelahnya akan menjadi tanggung jawab kami.<br><br>
        Demikian disampaikan, atas perhatian dan kerjasama yang baik diucapkan terima kasih.
        """
        try:
            send_email_with_attachments(req.to_email, subject, html_body, [pdf_path, excel_path])
        except Exception as e:
            return {
                "error": f"Failed to send email: {str(e)}. Please check your SMTP_USERNAME and SMTP_PASSWORD in .env.",
                "excel_path": excel_path,
                "pdf_path": pdf_path
            }

    return {
        "message": "Reports generated and email sent successfully!" if req.send_email else "Reports generated",
        "excel_path": excel_path,
        "pdf_path": pdf_path
    }
